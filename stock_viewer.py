import os
import sys
import time
import json
import math
import threading
import subprocess
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, colorchooser
import openpyxl
from openpyxl.styles import Font as XFont, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import yfinance as yf

# ── Defaults & preferences ───────────────────────────────────────
DEFAULTS = {
    "C_BG":        "#0d1b2a",
    "C_SIDEBAR":   "#112236",
    "C_PANEL":     "#1a2d42",
    "C_ACCENT":    "#1565c0",
    "C_ACCENT_LT": "#1e88e5",
    "C_TEAL":      "#00acc1",
    "C_TEXT":      "#dce8f5",
    "C_MUTED":     "#6a8aaa",
    "C_WHITE":     "#ffffff",
    "C_BORDER":    "#1e3450",
    "C_RPT_BG":    "#f4f7fb",
    "C_GREEN":     "#2e7d32",
    "C_RED":       "#c62828",
    "FONT":        "Segoe UI",
    "FONT_SZ":     10,
}
P = dict(DEFAULTS)   # live prefs — mutated by customize dialog

FIELD_GROUPS = {
    "Company Info":  ["Company Name", "Company Description"],
    "Market Data":   ["Mkt Cap/Net Assets (M)", "Price", "Prev Close",
                      "Change %", "Day High", "Day Low", "EPS (TTM)"],
    "Revenue":       ["Revenue TTM (M)", "Revenue 1Y (M)",
                      "Revenue 3Y (M)", "Revenue 5Y (M)"],
    "Net Income":    ["Net Income TTM (M)", "Net Income 1Y (M)",
                      "Net Income 3Y (M)", "Net Income 5Y (M)"],
    "Balance Sheet": ["Total Debt (M)", "Cash (M)"],
    "Other":         ["Last Updated"],
}

PRICE_FIELDS   = {"Price", "Prev Close", "Day High", "Day Low", "EPS (TTM)"}
MILLION_FIELDS = {
    "Mkt Cap/Net Assets (M)",
    "Revenue TTM (M)", "Revenue 1Y (M)", "Revenue 3Y (M)", "Revenue 5Y (M)",
    "Net Income TTM (M)", "Net Income 1Y (M)", "Net Income 3Y (M)", "Net Income 5Y (M)",
    "Total Debt (M)", "Cash (M)",
}

EXCEL_FIELDS = [
    "Stock Ticker", "Company Name", "Company Description",
    "Mkt Cap/Net Assets (M)", "Price", "Prev Close", "Change %",
    "Day High", "Day Low", "EPS (TTM)",
    "Revenue TTM (M)", "Revenue 1Y (M)", "Revenue 3Y (M)", "Revenue 5Y (M)",
    "Net Income TTM (M)", "Net Income 1Y (M)", "Net Income 3Y (M)", "Net Income 5Y (M)",
    "Total Debt (M)", "Cash (M)", "Last Updated",
]


# ── Helpers ──────────────────────────────────────────────────────
_WIN = sys.platform == "win32"


def _git(*cmd, **kwargs):
    """Run a git command; suppress console popup on Windows."""
    if _WIN:
        kwargs.setdefault("creationflags", subprocess.CREATE_NO_WINDOW)
    return subprocess.run(list(cmd), **kwargs)


def git_push_excel(filepath, ticker=""):
    """Commit and push the xlsx to GitHub after each save."""
    repo_dir = os.path.dirname(os.path.abspath(filepath))
    fname    = os.path.basename(filepath)
    msg      = f"update {ticker}" if ticker else "update stocks"
    try:
        _git("git", "-C", repo_dir, "add", fname,
             check=True, capture_output=True)
        result = _git("git", "-C", repo_dir, "commit", "-m", msg,
                      capture_output=True, text=True)
        if result.returncode != 0:
            if "nothing to commit" in result.stdout or "nothing to commit" in result.stderr:
                return
            return
        _git("git", "-C", repo_dir, "push",
             capture_output=True, text=True)
    except (FileNotFoundError, subprocess.CalledProcessError):
        pass


def get_base_dir():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def prefs_path():
    return os.path.join(get_base_dir(), "viewer_prefs.json")


def load_prefs():
    try:
        with open(prefs_path()) as f:
            P.update(json.load(f))
    except Exception:
        pass


def save_prefs():
    try:
        with open(prefs_path(), "w") as f:
            json.dump(P, f, indent=2)
    except Exception:
        pass


def is_na(raw):
    """Return True if the value is missing, NaN, or empty."""
    if raw is None:
        return True
    if isinstance(raw, float) and math.isnan(raw):
        return True
    s = str(raw).strip().lower()
    return s in ("", "n/a", "none", "nan", "nat")


def fmt_value(field, raw):
    """Return (display_string, color_tag) for a field value."""
    if is_na(raw):
        return "N/A", "val_na"
    if field == "Change %":
        try:
            v = float(raw)
            s = f"+{v:,.2f}%" if v >= 0 else f"{v:,.2f}%"
            return s, ("val_green" if v >= 0 else "val_red")
        except Exception:
            return "N/A", "val_na"
    if field in MILLION_FIELDS:
        try:
            v = float(raw)
            if math.isnan(v):
                return "N/A", "val_na"
            return f"${v:,.2f} M", "val"
        except Exception:
            return "N/A", "val_na"
    if field in PRICE_FIELDS:
        try:
            v = float(raw)
            if math.isnan(v):
                return "N/A", "val_na"
            return f"${v:,.2f}", "val"
        except Exception:
            return "N/A", "val_na"
    return str(raw), "val"


def fmt_price(raw):
    try:
        v = float(raw)
        return f"${v:,.2f}"
    except Exception:
        return str(raw)


# ── Excel I/O ────────────────────────────────────────────────────
def load_excel(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            data[str(row[0])] = {
                headers[i]: row[i]
                for i in range(min(len(headers), len(row)))
            }
    return data


def write_excel_header(ws):
    fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fnt  = XFont(bold=True, color="FFFFFF", size=11)
    col_w = {
        "Stock Ticker": 13, "Company Name": 28, "Company Description": 55,
        "Mkt Cap/Net Assets (M)": 22, "Price": 10, "Prev Close": 12,
        "Change %": 10, "Day High": 10, "Day Low": 10, "EPS (TTM)": 12,
        "Revenue TTM (M)": 18, "Revenue 1Y (M)": 16, "Revenue 3Y (M)": 16,
        "Revenue 5Y (M)": 16, "Net Income TTM (M)": 20, "Net Income 1Y (M)": 18,
        "Net Income 3Y (M)": 18, "Net Income 5Y (M)": 18,
        "Total Debt (M)": 15, "Cash (M)": 12, "Last Updated": 20,
    }
    for col, field in enumerate(EXCEL_FIELDS, 1):
        cell = ws.cell(row=1, column=col, value=field)
        cell.fill = fill
        cell.font = fnt
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = col_w.get(field, 14)
    ws.row_dimensions[1].height = 35


def save_to_excel(data, filepath):
    if os.path.exists(filepath):
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        # Read all existing rows into a dict keyed by ticker
        all_rows = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                h_len = min(len(headers), len(row))
                all_rows[str(row[0])] = {headers[i]: row[i] for i in range(h_len)}
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stock Data"
        write_excel_header(ws)
        all_rows = {}

    # Update or add the new ticker
    all_rows[data["Stock Ticker"]] = data

    # Sort alphabetically (case-insensitive)
    sorted_tickers = sorted(all_rows.keys(), key=str.upper)

    # Erase all data rows and rewrite in sorted order
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for i, ticker in enumerate(sorted_tickers):
        row_num = i + 2
        even    = row_num % 2 == 0
        fill    = PatternFill(
            start_color="D9E1F2" if even else "FFFFFF",
            end_color="D9E1F2" if even else "FFFFFF",
            fill_type="solid",
        )
        row_data = all_rows[ticker]
        for col, field in enumerate(EXCEL_FIELDS, 1):
            cell = ws.cell(row=row_num, column=col, value=row_data.get(field))
            cell.fill = fill
            cell.alignment = Alignment(vertical="top",
                                       wrap_text=(field == "Company Description"))
    wb.save(filepath)


def delete_from_excel(ticker, filepath):
    """Remove a ticker row from the Excel file, keeping all others sorted."""
    if not os.path.exists(filepath):
        return
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    all_rows = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and str(row[0]) != ticker:
            h_len = min(len(headers), len(row))
            all_rows[str(row[0])] = {headers[i]: row[i] for i in range(h_len)}

    sorted_tickers = sorted(all_rows.keys(), key=str.upper)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for i, t in enumerate(sorted_tickers):
        row_num = i + 2
        even    = row_num % 2 == 0
        fill    = PatternFill(
            start_color="D9E1F2" if even else "FFFFFF",
            end_color="D9E1F2" if even else "FFFFFF",
            fill_type="solid",
        )
        row_data = all_rows[t]
        for col, field in enumerate(EXCEL_FIELDS, 1):
            cell = ws.cell(row=row_num, column=col, value=row_data.get(field))
            cell.fill = fill
            cell.alignment = Alignment(vertical="top",
                                       wrap_text=(field == "Company Description"))
    wb.save(filepath)


# ── yfinance fetch ───────────────────────────────────────────────
def safe_m(v):
    try:
        f = float(v)
        return "N/A" if math.isnan(f) else round(f / 1e6, 2)
    except Exception:
        return "N/A"


def safe_v(v):
    try:
        f = float(v)
        return "N/A" if math.isnan(f) else round(f, 4)
    except Exception:
        return "N/A"


# Row-name variants used by yfinance across different regions / versions
_REV_ROWS = ["Total Revenue", "Operating Revenue", "TotalRevenue",
             "Net Interest Income", "Revenue"]
_NI_ROWS  = ["Net Income", "Net Income Common Stockholders",
             "Net Income From Continuing Operations", "NetIncome",
             "Net Income Including Noncontrolling Interests"]


def get_annual(fin, row_names, idx):
    """Try multiple row-name variants in the financials DataFrame."""
    if isinstance(row_names, str):
        row_names = [row_names]
    try:
        if fin is not None and not fin.empty and fin.shape[1] > idx:
            for row in row_names:
                if row in fin.index:
                    return safe_m(fin.iloc[:, idx].get(row))
    except Exception:
        pass
    return "N/A"


def build_description(info):
    d = info.get("longBusinessSummary")
    if d:
        return d
    qt = info.get("quoteType", "")
    parts = []
    if qt in ("ETF", "MUTUALFUND"):
        parts.append(f"{info.get('longName', '')} is an {qt}.")
        if info.get("category"):
            parts.append(f"Category: {info['category']}.")
        if info.get("fundFamily"):
            parts.append(f"Fund family: {info['fundFamily']}.")
    return " ".join(parts) if parts else "N/A"


def fetch_stock_data(symbol, retries=4, base_delay=5):
    for attempt in range(retries):
        try:
            t = yf.Ticker(symbol)
            info = t.info
            if not info or (info.get("currentPrice") is None
                            and info.get("regularMarketPrice") is None
                            and info.get("previousClose") is None):
                return None, f"No data found for '{symbol}'."

            price      = safe_v(info.get("currentPrice") or info.get("regularMarketPrice"))
            prev_close = safe_v(info.get("previousClose") or info.get("regularMarketPreviousClose"))
            try:
                change = round(((float(price) - float(prev_close)) / float(prev_close)) * 100, 2) \
                         if price != "N/A" and prev_close != "N/A" else "N/A"
            except Exception:
                change = "N/A"

            # Income statement — try newer API first, fall back to older
            fin = None
            for attr in ("income_stmt", "financials"):
                try:
                    f = getattr(t, attr)
                    if f is not None and not f.empty:
                        fin = f
                        break
                except Exception:
                    pass

            # Balance sheet for debt / cash fallback
            bs = None
            for attr in ("balance_sheet", "quarterly_balance_sheet"):
                try:
                    b = getattr(t, attr)
                    if b is not None and not b.empty:
                        bs = b
                        break
                except Exception:
                    pass

            # EPS — with calculated fallback
            eps = safe_v(info.get("trailingEps"))
            if eps == "N/A":
                try:
                    ni = info.get("netIncomeToCommon")
                    sh = info.get("sharesOutstanding")
                    if ni and sh:
                        eps = round(float(ni) / float(sh), 4)
                except Exception:
                    pass

            # Debt — with balance sheet fallback
            total_debt = safe_m(info.get("totalDebt"))
            if total_debt == "N/A" and bs is not None and not bs.empty:
                for rn in ("Total Debt", "TotalDebt", "Net Debt", "Long Term Debt"):
                    if rn in bs.index:
                        v = safe_m(bs.iloc[:, 0].get(rn))
                        if v != "N/A":
                            total_debt = v
                            break

            # Cash — with balance sheet fallback
            cash = safe_m(info.get("totalCash") or info.get("cash"))
            if cash == "N/A" and bs is not None and not bs.empty:
                for rn in ("Cash And Cash Equivalents", "Cash",
                           "Cash Cash Equivalents And Short Term Investments"):
                    if rn in bs.index:
                        v = safe_m(bs.iloc[:, 0].get(rn))
                        if v != "N/A":
                            cash = v
                            break

            return {
                "Stock Ticker":           symbol.upper(),
                "Company Name":           info.get("longName", "N/A"),
                "Company Description":    build_description(info),
                "Mkt Cap/Net Assets (M)": safe_m(info.get("marketCap") or info.get("totalAssets")),
                "Price":                  price,
                "Prev Close":             prev_close,
                "Change %":               change,
                "Day High":               safe_v(info.get("dayHigh") or info.get("regularMarketDayHigh")),
                "Day Low":                safe_v(info.get("dayLow") or info.get("regularMarketDayLow")),
                "EPS (TTM)":              eps,
                "Revenue TTM (M)":        safe_m(info.get("totalRevenue")),
                "Revenue 1Y (M)":         get_annual(fin, _REV_ROWS, 0),
                "Revenue 3Y (M)":         get_annual(fin, _REV_ROWS, 2),
                "Revenue 5Y (M)":         get_annual(fin, _REV_ROWS, 4),
                "Net Income TTM (M)":     safe_m(info.get("netIncomeToCommon")),
                "Net Income 1Y (M)":      get_annual(fin, _NI_ROWS, 0),
                "Net Income 3Y (M)":      get_annual(fin, _NI_ROWS, 2),
                "Net Income 5Y (M)":      get_annual(fin, _NI_ROWS, 4),
                "Total Debt (M)":         total_debt,
                "Cash (M)":               cash,
                "Last Updated":           datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }, None

        except Exception as e:
            err = str(e)
            if any(k in err.lower() for k in ("too many requests", "rate limit", "429")):
                wait = base_delay * (2 ** attempt)
                time.sleep(wait)
            else:
                return None, f"Error: {e}"
    return None, "Rate limited after retries. Wait a moment and try again."


# ── Add Ticker Dialog ─────────────────────────────────────────────
class AddTickerDialog(tk.Toplevel):
    def __init__(self, app):
        super().__init__(app)
        self.app = app
        self.title("Add Ticker")
        self.geometry("420x210")
        self.resizable(False, False)
        self.configure(bg=P["C_BG"])
        self.grab_set()
        self.transient(app)

        tk.Label(self, text="Add a New Ticker",
                 bg=P["C_BG"], fg=P["C_WHITE"],
                 font=(P["FONT"], 13, "bold")).pack(pady=(20, 4))
        tk.Label(self, text="Enter the ticker symbol to fetch and save:",
                 bg=P["C_BG"], fg=P["C_MUTED"],
                 font=(P["FONT"], 9)).pack()

        frm = tk.Frame(self, bg=P["C_BG"])
        frm.pack(pady=12)
        self.entry = tk.Entry(frm, font=(P["FONT"], 13),
                              bg=P["C_PANEL"], fg=P["C_WHITE"],
                              insertbackground=P["C_WHITE"],
                              relief="flat", width=14)
        self.entry.pack(side="left", ipady=6, padx=(0, 8))
        self.entry.bind("<Return>", lambda e: self._fetch())
        self.entry.focus()

        self.btn = tk.Button(frm, text="Fetch & Add",
                             bg=P["C_ACCENT_LT"], fg=P["C_WHITE"],
                             font=(P["FONT"], 10, "bold"),
                             relief="flat", cursor="hand2",
                             padx=10, pady=6,
                             command=self._fetch)
        self.btn.pack(side="left")

        self.status = tk.Label(self, text="",
                               bg=P["C_BG"], fg=P["C_TEAL"],
                               font=(P["FONT"], 9, "italic"))
        self.status.pack()

    def _fetch(self):
        sym = self.entry.get().strip().upper()
        if not sym:
            return
        if not self.app.filepath:
            messagebox.showerror("No File",
                                 "No Excel file is loaded. Open StockTracker first.",
                                 parent=self)
            return
        self.btn.config(state="disabled")
        self.status.config(text=f"Fetching {sym}…", fg=P["C_TEAL"])
        threading.Thread(target=self._do, args=(sym,), daemon=True).start()

    def _do(self, sym):
        data, err = fetch_stock_data(sym)
        if err:
            self.after(0, lambda: self.status.config(text=f"✗  {err}", fg=P["C_RED"]))
            self.after(0, lambda: self.btn.config(state="normal"))
            return
        save_to_excel(data, self.app.filepath)
        git_push_excel(self.app.filepath, ticker=sym)
        self.after(0, lambda: self._done(sym, data))

    def _done(self, sym, data):
        self.app.stock_data[sym] = data
        tickers = sorted(self.app.stock_data.keys())
        self.app.combo["values"] = tickers
        self.app.combo.set(sym)
        self.status.config(text=f"✓  {sym} added and saved!", fg=P["C_GREEN"])
        self.after(1800, self.destroy)


# ── Customize Dialog ──────────────────────────────────────────────
class CustomizeDialog(tk.Toplevel):
    COLOR_OPTS = [
        ("Background",        "C_BG"),
        ("Sidebar",           "C_SIDEBAR"),
        ("Accent",            "C_ACCENT"),
        ("Accent (light)",    "C_ACCENT_LT"),
        ("Highlight",         "C_TEAL"),
        ("Report Background", "C_RPT_BG"),
        ("Text",              "C_TEXT"),
        ("Positive (green)",  "C_GREEN"),
        ("Negative (red)",    "C_RED"),
    ]
    FONT_OPTIONS = [
        "Segoe UI", "Arial", "Calibri", "Helvetica",
        "Verdana", "Tahoma", "Georgia", "Courier New",
    ]

    def __init__(self, app):
        super().__init__(app)
        self.app        = app
        self.draft      = dict(P)
        self.hex_vars   = {}
        self.swatch_btns = {}
        self.title("Customize Layout")
        self.geometry("530x660")
        self.resizable(False, True)
        self.configure(bg=P["C_BG"])
        self.grab_set()
        self.transient(app)
        self._build()

    def _build(self):
        # ── Fixed header ──────────────────────────────────────
        hdr = tk.Frame(self, bg=P["C_BG"])
        hdr.pack(fill="x", padx=20, pady=(16, 0))
        tk.Label(hdr, text="Customize Layout",
                 bg=P["C_BG"], fg=P["C_WHITE"],
                 font=(P["FONT"], 14, "bold")).pack()
        tk.Label(hdr, text="Click a color swatch to change it, then press Apply.",
                 bg=P["C_BG"], fg=P["C_MUTED"],
                 font=(P["FONT"], 9)).pack(pady=(2, 10))

        # ── Fixed bottom buttons (packed BEFORE scrollable area) ──
        bf = tk.Frame(self, bg=P["C_BG"])
        bf.pack(side="bottom", fill="x", pady=14, padx=20)
        for text, cmd, bg_col in [
            ("✔  Apply",          self._apply,  P["C_ACCENT_LT"]),
            ("↺  Reset Defaults", self._reset,  "#555555"),
            ("✕  Cancel",         self.destroy, P["C_PANEL"]),
        ]:
            tk.Button(bf, text=text, command=cmd,
                      bg=bg_col, fg=P["C_WHITE"],
                      font=(P["FONT"], 10, "bold"),
                      relief="flat", cursor="hand2",
                      padx=14, pady=8
                      ).pack(side="left", padx=(0, 8))

        tk.Frame(self, bg=P["C_BORDER"], height=1).pack(
            side="bottom", fill="x", padx=20)

        # ── Scrollable middle section ──────────────────────────
        cv  = tk.Canvas(self, bg=P["C_BG"], highlightthickness=0)
        vsb = ttk.Scrollbar(self, orient="vertical", command=cv.yview)
        inner = tk.Frame(cv, bg=P["C_BG"])
        inner.bind("<Configure>",
                   lambda e: cv.configure(scrollregion=cv.bbox("all")))
        cv.create_window((0, 0), window=inner, anchor="nw")
        cv.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        cv.pack(fill="both", expand=True, padx=20)
        cv.bind("<MouseWheel>",
                lambda e: cv.yview_scroll(int(-1*(e.delta/120)), "units"))
        inner.bind("<MouseWheel>",
                   lambda e: cv.yview_scroll(int(-1*(e.delta/120)), "units"))

        # ── Colors section ────────────────────────────────────
        tk.Label(inner, text=" COLORS ", bg=P["C_BG"], fg=P["C_TEAL"],
                 font=(P["FONT"], 9, "bold")).pack(anchor="w", pady=(4, 2))
        tk.Frame(inner, bg=P["C_TEAL"], height=1).pack(fill="x", pady=(0, 6))

        for label, key in self.COLOR_OPTS:
            row = tk.Frame(inner, bg=P["C_PANEL"], pady=1)
            row.pack(fill="x", pady=2)

            tk.Label(row, text=label, bg=P["C_PANEL"], fg=P["C_TEXT"],
                     font=(P["FONT"], 9), width=20, anchor="w",
                     padx=10).pack(side="left")

            swatch = tk.Button(row, bg=self.draft[key],
                               relief="flat", width=7, cursor="hand2",
                               command=lambda k=key: self._pick(k))
            swatch.pack(side="left", padx=6, ipady=7)
            self.swatch_btns[key] = swatch

            hv = tk.StringVar(value=self.draft[key])
            self.hex_vars[key] = hv
            tk.Label(row, textvariable=hv, bg=P["C_PANEL"], fg="#a0bcd8",
                     font=("Courier New", 9), width=9).pack(side="left")

        # ── Font section ──────────────────────────────────────
        tk.Label(inner, text=" FONT ", bg=P["C_BG"], fg=P["C_TEAL"],
                 font=(P["FONT"], 9, "bold")).pack(anchor="w", pady=(14, 2))
        tk.Frame(inner, bg=P["C_TEAL"], height=1).pack(fill="x", pady=(0, 6))

        font_row = tk.Frame(inner, bg=P["C_PANEL"], pady=6)
        font_row.pack(fill="x", pady=2)

        tk.Label(font_row, text="Font Family", bg=P["C_PANEL"], fg=P["C_TEXT"],
                 font=(P["FONT"], 9), padx=10, width=14,
                 anchor="w").pack(side="left")

        self.font_var = tk.StringVar(value=self.draft["FONT"])
        # Plain OptionMenu so text is always visible on any bg
        om = tk.OptionMenu(font_row, self.font_var, *self.FONT_OPTIONS)
        om.config(bg=P["C_ACCENT"], fg=P["C_WHITE"],
                  activebackground=P["C_ACCENT_LT"], activeforeground=P["C_WHITE"],
                  font=(P["FONT"], 9), relief="flat",
                  highlightthickness=0, cursor="hand2", width=14)
        om["menu"].config(bg=P["C_PANEL"], fg=P["C_WHITE"],
                          activebackground=P["C_ACCENT"],
                          font=(P["FONT"], 9))
        om.pack(side="left", padx=6)

        tk.Label(font_row, text="Size", bg=P["C_PANEL"], fg=P["C_TEXT"],
                 font=(P["FONT"], 9), padx=6).pack(side="left")
        self.size_var = tk.IntVar(value=self.draft["FONT_SZ"])
        sp = tk.Spinbox(font_row, from_=8, to=16, textvariable=self.size_var,
                        width=4, font=(P["FONT"], 10),
                        bg=P["C_BG"], fg=P["C_WHITE"],
                        buttonbackground=P["C_BORDER"],
                        insertbackground=P["C_WHITE"], relief="flat")
        sp.pack(side="left", padx=6)

        tk.Label(inner, bg=P["C_BG"], height=1).pack()  # bottom spacer

    def _pick(self, key):
        result = colorchooser.askcolor(color=self.draft[key],
                                       title=f"Choose — {key}", parent=self)
        if result and result[1]:
            col = result[1]
            self.draft[key] = col
            self.swatch_btns[key].config(bg=col)
            self.hex_vars[key].set(col)

    def _apply(self):
        P.update(self.draft)
        P["FONT"]    = self.font_var.get()
        P["FONT_SZ"] = self.size_var.get()
        save_prefs()
        self.app._rebuild()
        self.destroy()

    def _reset(self):
        self.draft = dict(DEFAULTS)
        for key, btn in self.swatch_btns.items():
            btn.config(bg=self.draft[key])
            self.hex_vars[key].set(self.draft[key])
        self.font_var.set(DEFAULTS["FONT"])
        self.size_var.set(DEFAULTS["FONT_SZ"])


# ── Main App ─────────────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        load_prefs()
        self.title("Stock Report Viewer")
        self.geometry("1280x800")
        self.minsize(960, 620)
        self.configure(bg=P["C_BG"])
        self.stock_data = {}
        self.filepath   = None
        self.check_vars = {}
        self._style()
        self._header()
        self.body = None
        self._build_body()
        self._auto_load()

    def _style(self):
        s = ttk.Style(self)
        s.theme_use("clam")
        s.configure("TCombobox",
                    fieldbackground=P["C_PANEL"], background=P["C_PANEL"],
                    foreground=P["C_WHITE"], selectbackground=P["C_ACCENT"],
                    arrowcolor=P["C_TEAL"], bordercolor=P["C_BORDER"],
                    font=(P["FONT"], 11))
        s.map("TCombobox", fieldbackground=[("readonly", P["C_PANEL"])],
              foreground=[("readonly", P["C_WHITE"])])
        s.configure("Vertical.TScrollbar",
                    background=P["C_PANEL"], troughcolor=P["C_BG"],
                    arrowcolor=P["C_MUTED"])

    def _header(self):
        if hasattr(self, "_hdr") and self._hdr:
            self._hdr.destroy()
        h = tk.Frame(self, bg=P["C_ACCENT"], height=62)
        h.pack(fill="x", side="top")
        h.pack_propagate(False)
        self._hdr = h

        tk.Label(h, text="📊  Stock Report Viewer",
                 bg=P["C_ACCENT"], fg=P["C_WHITE"],
                 font=(P["FONT"], 17, "bold"), padx=22).pack(side="left", pady=14)

        # Right side controls
        frm = tk.Frame(h, bg=P["C_ACCENT"])
        frm.pack(side="right", padx=14, pady=10)

        for text, cmd, bg in [
            ("🎨 Customize", lambda: CustomizeDialog(self), "#0d47a1"),
            ("Browse…",     self._browse,                   "#0d47a1"),
        ]:
            tk.Button(frm, text=text, command=cmd,
                      bg=bg, fg=P["C_WHITE"],
                      font=(P["FONT"], 9), relief="flat",
                      cursor="hand2", padx=10, pady=4
                      ).pack(side="right", padx=3)

        self.file_lbl = tk.Label(frm, text="No file loaded",
                                  bg=P["C_ACCENT"], fg=P["C_WHITE"],
                                  font=(P["FONT"], 9, "italic"))
        self.file_lbl.pack(side="right", padx=(0, 10))
        tk.Label(frm, text="File:", bg=P["C_ACCENT"], fg=P["C_TEXT"],
                 font=(P["FONT"], 9)).pack(side="right")

    def _rebuild(self):
        """Rebuild body after theme change."""
        self._style()
        self._header()
        tickers  = list(self.stock_data.keys())
        sel_tick = self.combo.get() if hasattr(self, "combo") else ""
        sel_chks = {f: v.get() for f, v in self.check_vars.items()}
        old_path = self.filepath
        old_data = self.stock_data
        if self.body:
            self.body.destroy()
        self.check_vars = {}
        self._build_body()
        self.stock_data = old_data
        self.filepath   = old_path
        self.combo["values"] = tickers
        if sel_tick in tickers:
            self.combo.set(sel_tick)
        for f, v in sel_chks.items():
            if f in self.check_vars:
                self.check_vars[f].set(v)
        if self.filepath:
            self.file_lbl.config(text=os.path.basename(self.filepath))
        self.configure(bg=P["C_BG"])

    def _build_body(self):
        body = tk.Frame(self, bg=P["C_BG"])
        body.pack(fill="both", expand=True)
        self.body = body
        self._sidebar(body)
        tk.Frame(body, bg=P["C_BORDER"], width=1).pack(side="left", fill="y")
        self._report_pane(body)

    # ── Auto-load ─────────────────────────────────────────────
    def _auto_load(self):
        p = os.path.join(get_base_dir(), "stocks.xlsx")
        if os.path.exists(p):
            self._load(p)

    def _load(self, path):
        try:
            self.stock_data = load_excel(path)
            self.filepath   = path
            tickers = sorted(self.stock_data.keys())
            self.combo["values"] = tickers
            if tickers:
                self.combo.set(tickers[0])
            self.file_lbl.config(text=os.path.basename(path))
        except Exception as e:
            messagebox.showerror("Error", f"Could not load:\n{e}")

    def _browse(self):
        p = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if p:
            self._load(p)

    # ── Sidebar ───────────────────────────────────────────────
    def _sidebar(self, parent):
        sb = tk.Frame(parent, bg=P["C_SIDEBAR"], width=268)
        sb.pack(side="left", fill="y")
        sb.pack_propagate(False)

        def lbl(text, fg, sz=8):
            return tk.Label(sb, text=text, bg=P["C_SIDEBAR"], fg=fg,
                            font=(P["FONT"], sz, "bold"))

        lbl("TICKER", P["C_MUTED"]).pack(anchor="w", padx=18, pady=(20, 2))
        self.combo = ttk.Combobox(sb, font=(P["FONT"], 11),
                                   state="readonly", width=22)
        self.combo.pack(fill="x", padx=18)

        tk.Frame(sb, bg=P["C_BORDER"], height=1).pack(fill="x", padx=18, pady=14)
        lbl("FIELDS TO DISPLAY", P["C_MUTED"]).pack(anchor="w", padx=18)

        brow = tk.Frame(sb, bg=P["C_SIDEBAR"])
        brow.pack(fill="x", padx=18, pady=(6, 8))
        for txt, state in [("Select All", True), ("Clear All", False)]:
            tk.Button(brow, text=txt,
                      command=lambda s=state: [v.set(s) for v in self.check_vars.values()],
                      bg=P["C_PANEL"], fg=P["C_TEXT"],
                      font=(P["FONT"], 8), relief="flat",
                      cursor="hand2", padx=6, pady=2).pack(side="left", padx=(0, 4))

        # Scrollable checkboxes
        wrap = tk.Frame(sb, bg=P["C_SIDEBAR"])
        wrap.pack(fill="both", expand=True)
        cv  = tk.Canvas(wrap, bg=P["C_SIDEBAR"], highlightthickness=0)
        vsb = ttk.Scrollbar(wrap, orient="vertical", command=cv.yview)
        inner = tk.Frame(cv, bg=P["C_SIDEBAR"])
        inner.bind("<Configure>",
                   lambda e: cv.configure(scrollregion=cv.bbox("all")))
        cv.create_window((0, 0), window=inner, anchor="nw")
        cv.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        cv.pack(side="left", fill="both", expand=True, padx=(14, 0))

        def scroll(e):
            cv.yview_scroll(int(-1 * (e.delta / 120)), "units")
        cv.bind("<MouseWheel>", scroll)
        inner.bind("<MouseWheel>", scroll)

        for group, fields in FIELD_GROUPS.items():
            tk.Label(inner, text=group.upper(),
                     bg=P["C_SIDEBAR"], fg=P["C_TEAL"],
                     font=(P["FONT"], 7, "bold"), pady=5).pack(anchor="w")
            for f in fields:
                var = tk.BooleanVar(value=True)
                self.check_vars[f] = var
                cb = tk.Checkbutton(inner, text=f, variable=var,
                                    bg=P["C_SIDEBAR"], fg=P["C_TEXT"],
                                    selectcolor=P["C_PANEL"],
                                    activebackground=P["C_SIDEBAR"],
                                    activeforeground=P["C_TEXT"],
                                    font=(P["FONT"], 9), anchor="w", padx=4)
                cb.pack(fill="x")
                cb.bind("<MouseWheel>", scroll)

        tk.Frame(sb, bg=P["C_BORDER"], height=1).pack(fill="x", padx=18, pady=10)

        # Push to GitHub button
        tk.Button(sb, text="↑  Push to GitHub",
                  bg="#2e7d32", fg=P["C_WHITE"],
                  font=(P["FONT"], 10, "bold"),
                  relief="flat", cursor="hand2",
                  padx=10, pady=8,
                  activebackground="#388e3c",
                  command=self._push_to_github
                  ).pack(fill="x", padx=18, pady=(0, 6))

        # Sync from GitHub button
        tk.Button(sb, text="↓  Sync from GitHub",
                  bg=P["C_ACCENT"], fg=P["C_WHITE"],
                  font=(P["FONT"], 10, "bold"),
                  relief="flat", cursor="hand2",
                  padx=10, pady=8,
                  activebackground=P["C_ACCENT_LT"],
                  command=self._sync_from_github
                  ).pack(fill="x", padx=18, pady=(0, 6))

        # Add Ticker button
        tk.Button(sb, text="＋  Add Another Ticker",
                  bg=P["C_TEAL"], fg=P["C_WHITE"],
                  font=(P["FONT"], 10, "bold"),
                  relief="flat", cursor="hand2",
                  padx=10, pady=8,
                  activebackground="#0097a7",
                  command=lambda: AddTickerDialog(self)
                  ).pack(fill="x", padx=18, pady=(0, 6))

        # Delete Ticker button
        tk.Button(sb, text="🗑  Delete Ticker",
                  bg=P["C_RED"], fg=P["C_WHITE"],
                  font=(P["FONT"], 10, "bold"),
                  relief="flat", cursor="hand2",
                  padx=10, pady=8,
                  activebackground="#b71c1c",
                  command=self._delete_ticker
                  ).pack(fill="x", padx=18, pady=(0, 6))

        # Generate button
        tk.Button(sb, text="▶   GENERATE REPORT",
                  bg=P["C_ACCENT_LT"], fg=P["C_WHITE"],
                  font=(P["FONT"], 11, "bold"),
                  relief="flat", cursor="hand2",
                  padx=10, pady=11,
                  activebackground="#1976d2",
                  command=self._generate
                  ).pack(fill="x", padx=18, pady=(0, 18))

    # ── Report pane ───────────────────────────────────────────
    def _report_pane(self, parent):
        pane = tk.Frame(parent, bg=P["C_RPT_BG"])
        pane.pack(side="left", fill="both", expand=True)

        self.rpt_title = tk.Label(
            pane,
            text="  Select a ticker and click  ▶  Generate Report",
            bg="#dce8f5", fg="#5a7a9a",
            font=(P["FONT"], 11, "italic"),
            anchor="w", padx=20, pady=13)
        self.rpt_title.pack(fill="x")

        frm = tk.Frame(pane, bg=P["C_RPT_BG"])
        frm.pack(fill="both", expand=True)

        self.txt = tk.Text(frm, bg=P["C_RPT_BG"], fg="#1a2332",
                           font=(P["FONT"], P["FONT_SZ"]),
                           wrap="word", relief="flat",
                           padx=36, pady=24,
                           state="disabled", cursor="arrow",
                           selectbackground=P["C_ACCENT_LT"])
        vsb2 = ttk.Scrollbar(frm, orient="vertical", command=self.txt.yview)
        self.txt.configure(yscrollcommand=vsb2.set)
        vsb2.pack(side="right", fill="y")
        self.txt.pack(side="left", fill="both", expand=True)
        self.txt.bind("<MouseWheel>",
                      lambda e: self.txt.yview_scroll(int(-1*(e.delta/120)), "units"))
        self._setup_tags()

    def _setup_tags(self):
        T = self.txt
        F = P["FONT"]
        Z = P["FONT_SZ"]
        T.tag_configure("ticker_sym",
            font=(F, 34, "bold"), foreground="#0d1b2a",
            spacing1=10, spacing3=2)
        T.tag_configure("co_name",
            font=(F, 14), foreground="#2a5080", spacing3=4)
        T.tag_configure("desc",
            font=(F, Z - 1), foreground="#5a6e85", spacing3=18)
        T.tag_configure("sec_hdr",
            font=(F, Z, "bold"), foreground=P["C_WHITE"],
            background=P["C_ACCENT"], spacing1=14, spacing3=6)
        T.tag_configure("lbl",
            font=(F, Z), foreground="#6a85a0")
        T.tag_configure("val",
            font=(F, Z, "bold"), foreground="#0d1b2a")
        T.tag_configure("val_green",
            font=(F, Z, "bold"), foreground=P["C_GREEN"])
        T.tag_configure("val_red",
            font=(F, Z, "bold"), foreground=P["C_RED"])
        T.tag_configure("val_na",
            font=(F, Z, "italic"), foreground=P["C_RED"])
        T.tag_configure("row0", background="#edf2fa",
            spacing1=5, spacing3=5)
        T.tag_configure("row1", background="#f8fafd",
            spacing1=5, spacing3=5)
        T.tag_configure("foot",
            font=(F, Z - 1, "italic"), foreground="#9ab0c5", spacing1=18)

    # ── Push to GitHub ────────────────────────────────────────
    def _push_to_github(self):
        if not self.filepath:
            messagebox.showerror("No File", "No Excel file is loaded.")
            return
        try:
            git_push_excel(self.filepath, ticker="manual push")
            messagebox.showinfo(
                "Pushed",
                "stocks.xlsx pushed to GitHub.\n"
                "Press Refresh Data on the web to see changes.")
        except Exception as e:
            messagebox.showerror("Error", f"Push failed:\n{e}")

    # ── Sync from GitHub ──────────────────────────────────────
    def _sync_from_github(self):
        if not self.filepath:
            messagebox.showerror("No File", "No Excel file is loaded.")
            return
        try:
            repo_dir = os.path.dirname(os.path.abspath(self.filepath))
            fname    = os.path.basename(self.filepath)
            # Fetch latest from origin then checkout only stocks.xlsx
            _git("git", "-C", repo_dir, "fetch", "origin",
                 capture_output=True)
            r = _git("git", "-C", repo_dir, "checkout", "origin/main",
                     "--", fname, capture_output=True, text=True)
            if r.returncode == 0:
                self._load(self.filepath)
                messagebox.showinfo(
                    "Synced", f"'{fname}' updated from GitHub.\n"
                              f"{len(self.stock_data)} ticker(s) loaded.")
            else:
                messagebox.showerror(
                    "Sync Failed",
                    f"Could not sync from GitHub:\n{r.stderr[:300]}")
        except Exception as e:
            messagebox.showerror("Error", f"Sync error:\n{e}")

    # ── Delete Ticker ─────────────────────────────────────────
    def _delete_ticker(self):
        ticker = self.combo.get().strip()
        if not ticker:
            messagebox.showwarning("No Ticker", "Please select a ticker first.")
            return
        if not self.filepath:
            messagebox.showerror("No File", "No Excel file is loaded.")
            return
        if not messagebox.askyesno(
            "Confirm Delete",
            f"Delete '{ticker}' from the Excel file?\nThis cannot be undone.",
            parent=self
        ):
            return
        try:
            delete_from_excel(ticker, self.filepath)
            git_push_excel(self.filepath, ticker=f"delete {ticker}")
            del self.stock_data[ticker]
            tickers = sorted(self.stock_data.keys())
            self.combo["values"] = tickers
            if tickers:
                self.combo.set(tickers[0])
            else:
                self.combo.set("")
                self.txt.config(state="normal")
                self.txt.delete("1.0", "end")
                self.txt.config(state="disabled")
                self.rpt_title.config(
                    text="  Select a ticker and click  ▶  Generate Report",
                    bg="#dce8f5", fg="#5a7a9a",
                    font=(P["FONT"], 11, "italic"))
        except Exception as e:
            messagebox.showerror("Error", f"Could not delete {ticker}:\n{e}")

    # ── Report generation ─────────────────────────────────────
    def _generate(self):
        ticker = self.combo.get().strip()
        if not ticker:
            messagebox.showwarning("No Ticker", "Please select a ticker first.")
            return
        if ticker not in self.stock_data:
            messagebox.showerror("Not Found", f"'{ticker}' not found.")
            return

        rec = self.stock_data[ticker]
        sel = {f for f, v in self.check_vars.items() if v.get()}
        co  = rec.get("Company Name", ticker)

        self.rpt_title.config(
            text=f"  {ticker}   ·   {co}",
            bg=P["C_ACCENT"], fg=P["C_WHITE"],
            font=(P["FONT"], 11, "bold"))

        T = self.txt
        T.config(state="normal")
        T.delete("1.0", "end")

        # ── Header ────────────────────────────────────────
        T.insert("end", f"\n  {ticker}\n", "ticker_sym")
        if "Company Name" in sel:
            T.insert("end", f"  {co}\n", "co_name")
        if "Company Description" in sel:
            d = rec.get("Company Description") or "N/A"
            T.insert("end", f"  {d}\n\n", "desc")

        # ── Sections ──────────────────────────────────────
        def section(title, fields):
            visible = [f for f in fields
                       if f in sel and f not in ("Company Name", "Company Description")]
            if not visible:
                return
            T.insert("end", f"   {title.upper()}   \n", "sec_hdr")
            for i, f in enumerate(visible):
                raw   = rec.get(f)
                s, vt = fmt_value(f, raw)
                rt    = "row0" if i % 2 == 0 else "row1"
                T.insert("end", f"  {f:<34}", ("lbl", rt))
                T.insert("end", f"{s}\n", (vt, rt))

        section("Market Data", [
            "Mkt Cap/Net Assets (M)", "Price", "Prev Close",
            "Change %", "Day High", "Day Low", "EPS (TTM)"])
        section("Revenue", [
            "Revenue TTM (M)", "Revenue 1Y (M)",
            "Revenue 3Y (M)", "Revenue 5Y (M)"])
        section("Net Income", [
            "Net Income TTM (M)", "Net Income 1Y (M)",
            "Net Income 3Y (M)", "Net Income 5Y (M)"])
        section("Balance Sheet", ["Total Debt (M)", "Cash (M)"])

        if "Last Updated" in sel:
            T.insert("end",
                     f"\n  Last Updated:  {rec.get('Last Updated', 'N/A')}\n",
                     "foot")

        T.config(state="disabled")
        T.yview_moveto(0)


if __name__ == "__main__":
    App().mainloop()
