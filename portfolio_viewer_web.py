"""
Gaboch Portfolio Viewer — Web App (Streamlit)

Run:  streamlit run portfolio_viewer_web.py

Reads Gaboch_portfolio.xlsx (row-pair format) and displays active
stock transactions in a browser-based UI with theme customization.
"""
from __future__ import annotations

import base64
import contextlib
import hashlib
import io
import json
import math
import os
import re
import time
from datetime import datetime

import openpyxl
import requests
import streamlit as st
import yfinance as yf

# ─────────────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────────────
SHEET_NAME       = "Gaboch_portfolio"
DEFAULT_EXCEL    = r"C:\Gaboch_Portfolio\Gaboch_portfolio.xlsx"
_WEB_PREFS_FILE  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "web_prefs.json")
_GH_PREFS_FILE   = "web_prefs.json"

CANADIAN_EXCHANGES = {"XTSE", "XTSX", "XCNQ", "TSX", "TSXV"}

THEME_PRESETS: dict[str, dict] = {
    "Dark Navy": {
        "bg": "#0d1b2a", "card_bg": "#1b2a3b", "accent": "#1565c0",
        "text": "#e8edf2", "profit": "#4caf50", "loss": "#ef5350",
        "header": "#64b5f6", "subhead": "#90caf9",
        "border": "#1e3a5f", "font_size": 15,
    },
    "Steel Blue": {
        "bg": "#0a1929", "card_bg": "#102a43", "accent": "#0288d1",
        "text": "#dce8f0", "profit": "#26a69a", "loss": "#ef5350",
        "header": "#4fc3f7", "subhead": "#81d4fa",
        "border": "#0d47a1", "font_size": 15,
    },
    "Forest Green": {
        "bg": "#0a1a0a", "card_bg": "#0d2b0d", "accent": "#388e3c",
        "text": "#e0ebe0", "profit": "#66bb6a", "loss": "#ef5350",
        "header": "#81c784", "subhead": "#a5d6a7",
        "border": "#1b5e20", "font_size": 15,
    },
    "Deep Purple": {
        "bg": "#1a0533", "card_bg": "#220a40", "accent": "#7b1fa2",
        "text": "#ede0f5", "profit": "#ba68c8", "loss": "#ef5350",
        "header": "#ce93d8", "subhead": "#e1bee7",
        "border": "#4a148c", "font_size": 15,
    },
    "Light Mode": {
        "bg": "#f4f6f9", "card_bg": "#e8edf5", "accent": "#1565c0",
        "text": "#0d1117", "profit": "#2e7d32", "loss": "#c62828",
        "header": "#1565c0", "subhead": "#1565c0",
        "border": "#b0bec5", "font_size": 15,
    },
}

DEFAULTS = THEME_PRESETS["Dark Navy"]

# Maps theme color keys → their color-picker widget keys (mirrors stock_viewer pattern)
_CP_KEYS = {
    "profit": "cp_profit", "loss": "cp_loss",
    "header": "cp_header", "subhead": "cp_sub",
    "accent": "cp_accent", "text": "cp_text",
}


# ─────────────────────────────────────────────────────────────────────────────
# Data helpers
# ─────────────────────────────────────────────────────────────────────────────

def is_canadian(*fields) -> bool:
    """Return True if any field contains a Canadian exchange code, .TO/.V/.CN
    suffix, or an account name that mentions CAD."""
    for val in fields:
        s = str(val or "").strip().upper()
        for code in CANADIAN_EXCHANGES:
            if f"({code}:" in s or f"({code} " in s:
                return True
        if s.endswith(".TO") or s.endswith(".V") or s.endswith(".CN"):
            return True
        if " CAD " in f" {s} ":
            return True
    return False


def extract_company_name(ticker_full: str) -> str:
    m = re.match(r'^(.+?)\s*\(', str(ticker_full).strip())
    return m.group(1).strip() if m else str(ticker_full).strip()


def fmt_currency(value, canadian: bool, show_sign: bool = False) -> str:
    sym = "CAD$" if canadian else "USD$"
    if value is None:
        return "N/A"
    try:
        v = float(value)
        if show_sign:
            return f"+{sym} {v:,.2f}" if v >= 0 else f"-{sym} {abs(v):,.2f}"
        return f"{sym} {v:,.2f}"
    except (TypeError, ValueError):
        return f"{sym} {value}"


def fmt_pct(value) -> str:
    if value is None:
        return "N/A"
    try:
        v = float(value)
        return f"{'+' if v >= 0 else ''}{v:.2f}%"
    except (TypeError, ValueError):
        return str(value)


def fmt_shares(value) -> str:
    if value is None:
        return "N/A"
    try:
        v = float(value)
        return f"{int(v):,}" if v == int(v) else f"{v:,.4f}"
    except (TypeError, ValueError):
        return str(value)


def fmt_date(value) -> str:
    if value is None:
        return "—"
    if isinstance(value, str) and not value.strip():
        return "—"
    if isinstance(value, datetime):
        return value.strftime(f"%A, %B {value.day}, %Y")
    return str(value)


# ── Yahoo Finance helpers ─────────────────────────────────────────────────────

_TICKER_RE = re.compile(
    r'^[A-Z0-9]{1,6}(\.(TO|V|CN))?$'
    r'|^[A-Z0-9]{1,5}-(USD|CAD|GBP|EUR|BTC)$'
)


def _safe_v(v) -> float | None:
    """Safely convert to float; return None on NaN or error."""
    try:
        f = float(v)
        return None if math.isnan(f) else round(f, 4)
    except Exception:
        return None


def extract_yf_ticker(ticker_full: str) -> str:
    """Extract a yfinance-compatible symbol from 'NAME (EXCHANGE:TICKER)'."""
    s = str(ticker_full).strip()
    if _TICKER_RE.match(s.upper()):
        return s
    m = re.search(r'\(([^:)]+):([^)]+)\)', s)
    if not m:
        return ""
    exchange = m.group(1).strip().upper()
    symbol   = m.group(2).strip()
    if exchange in {"XTSE", "TSX"}:
        return symbol + ".TO"
    if exchange in {"XTSX", "TSXV"}:
        return symbol + ".V"
    if exchange in {"XCNQ"}:
        return symbol + ".CN"
    return symbol


def _build_desc(info: dict) -> str:
    """Build a company/ETF description from yfinance info (from stock_viewer_web)."""
    d = info.get("longBusinessSummary")
    if d:
        return d
    qt, parts = info.get("quoteType", ""), []
    if qt in ("ETF", "MUTUALFUND"):
        parts.append(f"{info.get('longName', '')} is an {qt}.")
        if info.get("category"):
            parts.append(f"Category: {info['category']}.")
        if info.get("fundFamily"):
            parts.append(f"Fund family: {info['fundFamily']}.")
    return " ".join(parts) if parts else ""


@st.cache_data(ttl=60)
def fetch_yf_info(yf_ticker: str, retries: int = 3, base_delay: float = 5.0) -> dict:
    """Return company name, description, live price, and currency from Yahoo Finance.

    Includes retry logic with exponential back-off on rate-limit errors
    (pattern from stock_viewer_web.py).
    """
    if not yf_ticker:
        return {}
    for attempt in range(retries):
        try:
            with contextlib.redirect_stderr(io.StringIO()), \
                 contextlib.redirect_stdout(io.StringIO()):
                t    = yf.Ticker(yf_ticker)
                info = t.info
            if not info:
                return {}
            price = _safe_v(
                info.get("currentPrice")
                or info.get("regularMarketPrice")
                or info.get("previousClose")
            )
            return {
                "company_name":  info.get("longName") or info.get("shortName") or "",
                "company_desc":  _build_desc(info),
                "current_price": price,
                "is_canadian":   info.get("currency", "USD").upper() == "CAD",
            }
        except Exception as exc:
            err_str = str(exc).lower()
            if any(k in err_str for k in ("too many requests", "rate limit", "429")):
                time.sleep(base_delay * (2 ** attempt))   # exponential back-off
            else:
                return {"company_name": "", "company_desc": "",
                        "current_price": None, "is_canadian": False}
    return {"company_name": "", "company_desc": "", "current_price": None, "is_canadian": False}


def _parse_rows(rows: list) -> dict:
    """Core row-pair parser shared by both load functions."""
    portfolio: dict = {}
    i = 0
    while i + 1 < len(rows):
        r1 = rows[i]      # purchase / even Excel row
        r2 = rows[i + 1]  # current  / odd  Excel row
        i += 2

        if all(c is None for c in r1):
            continue
        if r1[0] != 1:
            continue

        # Columns C and D carry the full ticker name; use whichever is not an error
        c2 = str(r1[2] or "")
        d2 = str(r1[3] or "")
        c3 = str(r2[2] or "")
        d3 = str(r2[3] or "")
        ticker_full = next((v for v in [c2, d2] if v and not v.startswith("#")), c2)
        ticker      = next((v for v in [c3, d3, c2, d2] if v and not v.startswith("#")), "").strip()
        if not ticker:
            continue

        txn = {
            "account":        r1[1],   # B
            "current_price":  r1[4],   # E
            "purchase_price": r1[5],   # F (even row = purchase price)
            "shares":         r1[6],   # G
            "subtotal":       r1[7],   # H
            "action":         r1[8],   # I
            "profits_usd":    r1[9],   # J
            "profits_pct":    r1[10],  # K
            "purchase_date":  r2[11],  # L — dates live in the odd (simple-ticker) row
            "selling_date":   r2[12],  # M
        }

        if ticker not in portfolio:
            account_name = str(r1[1] or "")
            is_cad = is_canadian(c2, d2, c3, d3, account_name)
            yf_tk  = extract_yf_ticker(ticker_full)
            # If extraction failed but stock is Canadian, build from first word + .TO
            if not yf_tk and is_cad:
                base = ticker.split()[0] if ticker else ""
                if base and _TICKER_RE.match((base + ".TO").upper()):
                    yf_tk = base + ".TO"
            portfolio[ticker] = {
                "company_name": extract_company_name(ticker_full),
                "ticker_full":  ticker_full,
                "yf_ticker":    yf_tk,
                "is_canadian":  is_cad,
                "transactions": [],
            }
        portfolio[ticker]["transactions"].append(txn)

    return portfolio


@st.cache_data(ttl=60)
def load_portfolio(filepath: str) -> dict:
    """Load from a local file path (laptop). Cached 60 s."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    return _parse_rows(list(wb[SHEET_NAME].iter_rows(min_row=2, values_only=True)))


@st.cache_data(ttl=60)
def load_portfolio_bytes(data: bytes) -> dict:
    """Load from uploaded bytes (Streamlit Cloud). Cached 60 s."""
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    return _parse_rows(list(wb[SHEET_NAME].iter_rows(min_row=2, values_only=True)))


# ── GitHub integration (same pattern as stock_viewer_web) ────────────────────

_GH_FILENAME = "Gaboch_portfolio.xlsx"


def _gh_secrets() -> tuple[str, str, str]:
    """Return (token, owner, repo) from Streamlit secrets."""
    return (
        st.secrets["GITHUB_TOKEN"],
        st.secrets["GITHUB_OWNER"],
        st.secrets["GITHUB_REPO"],
    )


@st.cache_data(ttl=60, show_spinner=False)
def load_portfolio_github() -> tuple[dict, str | None]:
    """Fetch Gaboch_portfolio.xlsx from the GitHub repo. Returns (portfolio, error)."""
    try:
        token, owner, repo = _gh_secrets()
    except KeyError as exc:
        return {}, f"Missing Streamlit secret: {exc}"

    url = (
        f"https://api.github.com/repos/{owner}/{repo}"
        f"/contents/{_GH_FILENAME}"
    )
    headers = {
        "Authorization":        f"Bearer {token}",
        "Accept":               "application/vnd.github.raw+json",
        "X-GitHub-Api-Version": "2022-11-28",
        "Cache-Control":        "no-cache",
        "Pragma":               "no-cache",
    }
    try:
        r = requests.get(f"{url}?_={int(time.time())}", headers=headers, timeout=20)
        if r.status_code == 200:
            return load_portfolio_bytes(r.content), None
        if r.status_code == 404:
            return {}, "not_found"   # sentinel — file not in repo yet
        return {}, f"GitHub error {r.status_code}: {r.text[:200]}"
    except Exception as exc:
        return {}, f"Connection error: {exc}"


def push_excel_to_github(data: bytes) -> tuple[bool, str | None]:
    """Upload Gaboch_portfolio.xlsx to the GitHub repo (creates or updates)."""
    try:
        token, owner, repo = _gh_secrets()
    except KeyError as exc:
        return False, f"Missing secret: {exc}"

    url = (
        f"https://api.github.com/repos/{owner}/{repo}"
        f"/contents/{_GH_FILENAME}"
    )
    req_headers = {
        "Authorization":        f"Bearer {token}",
        "X-GitHub-Api-Version": "2022-11-28",
    }
    content = base64.b64encode(data).decode()

    # Fetch current SHA so we can update (not just create)
    r = requests.get(url, headers=req_headers, timeout=15)
    sha = r.json().get("sha", "") if r.status_code == 200 else ""

    payload: dict = {"message": f"update {_GH_FILENAME}", "content": content}
    if sha:
        payload["sha"] = sha

    for attempt in range(3):
        try:
            r = requests.put(url, headers=req_headers, json=payload, timeout=30)
            if r.status_code in (200, 201):
                return True, None
            if r.status_code == 409:          # SHA conflict — retry
                time.sleep(1)
                r2 = requests.get(url, headers=req_headers, timeout=15)
                payload["sha"] = r2.json().get("sha", "")
                continue
            if r.status_code == 403:
                return False, (
                    "**403 Permission Denied** — your GitHub token is read-only.\n\n"
                    "Fix: GitHub → Settings → Developer settings → "
                    "Personal access tokens → create token with **`repo`** scope, "
                    "then update **GITHUB_TOKEN** in Streamlit → Manage app → Secrets."
                )
            return False, f"GitHub push failed ({r.status_code}): {r.text[:200]}"
        except Exception as exc:
            return False, f"Error: {exc}"
    return False, "Could not save after 3 attempts. Try again."


def _auto_push_if_changed(filepath: str) -> None:
    """Silently push local Excel to GitHub when the file content has changed.

    Uses an MD5 hash stored in session_state so the push only happens once
    per unique file version — not on every Streamlit rerun.
    """
    try:
        with open(filepath, "rb") as f:
            raw = f.read()
        current_hash = hashlib.md5(raw).hexdigest()
    except Exception:
        return
    if st.session_state.get("_excel_hash") == current_hash:
        return   # file unchanged since last push — skip
    ok, _ = push_excel_to_github(raw)
    if ok:
        st.session_state["_excel_hash"] = current_hash


# ─────────────────────────────────────────────────────────────────────────────
# Theme persistence
# ─────────────────────────────────────────────────────────────────────────────

def _load_theme_prefs() -> dict:
    """Load saved theme: local file first (laptop), then GitHub (cloud)."""
    # Laptop: local JSON file
    try:
        with open(_WEB_PREFS_FILE) as f:
            return json.load(f)
    except Exception:
        pass
    # Cloud: GitHub
    try:
        token, owner, repo = _gh_secrets()
        url = (f"https://api.github.com/repos/{owner}/{repo}"
               f"/contents/{_GH_PREFS_FILE}")
        r = requests.get(url, headers={
            "Authorization": f"Bearer {token}",
            "Accept": "application/vnd.github.raw+json",
            "X-GitHub-Api-Version": "2022-11-28",
        }, timeout=10)
        if r.status_code == 200:
            return json.loads(r.content)
    except Exception:
        pass
    return {}


def _save_theme_prefs(theme: dict, local_mode: bool) -> None:
    """Save theme to local file (laptop) or GitHub (cloud)."""
    if local_mode:
        try:
            with open(_WEB_PREFS_FILE, "w") as f:
                json.dump(theme, f, indent=2)
        except Exception:
            pass
    else:
        try:
            token, owner, repo = _gh_secrets()
            url = (f"https://api.github.com/repos/{owner}/{repo}"
                   f"/contents/{_GH_PREFS_FILE}")
            hdrs = {"Authorization": f"Bearer {token}",
                    "X-GitHub-Api-Version": "2022-11-28"}
            content = base64.b64encode(json.dumps(theme, indent=2).encode()).decode()
            r = requests.get(url, headers=hdrs, timeout=10)
            sha = r.json().get("sha", "") if r.status_code == 200 else ""
            payload: dict = {"message": "update web_prefs", "content": content}
            if sha:
                payload["sha"] = sha
            requests.put(url, headers=hdrs, json=payload, timeout=15)
        except Exception:
            pass


def _persist_theme_if_changed(theme: dict, local_mode: bool) -> None:
    """Save theme only when it has changed since last save (avoids redundant I/O)."""
    h = hashlib.md5(json.dumps(theme, sort_keys=True).encode()).hexdigest()
    if st.session_state.get("_theme_save_hash") == h:
        return
    _save_theme_prefs(theme, local_mode)
    st.session_state["_theme_save_hash"] = h


# ─────────────────────────────────────────────────────────────────────────────
# Theme / CSS
# ─────────────────────────────────────────────────────────────────────────────

def _init_theme() -> None:
    """Seed session state on first run — restores saved theme if available."""
    if "theme" not in st.session_state:
        saved = _load_theme_prefs()
        st.session_state.theme = {**DEFAULTS, **saved}
    t = st.session_state.theme
    # Seed color-picker widget keys from current theme so they stay in sync
    for tk, ck in _CP_KEYS.items():
        if ck not in st.session_state:
            st.session_state[ck] = t[tk]
    if "sl_font" not in st.session_state:
        st.session_state["sl_font"] = t["font_size"]


def _get_theme() -> dict:
    return st.session_state.theme


def _apply_preset(preset: dict) -> None:
    """Apply a preset and update both theme dict and color-picker widget keys."""
    st.session_state.theme = dict(preset)
    t = st.session_state.theme
    for tk, ck in _CP_KEYS.items():
        st.session_state[ck] = t[tk]
    st.session_state["sl_font"] = t["font_size"]


def _inject_css(t: dict) -> None:
    fs = t["font_size"]
    st.markdown(f"""
    <style>
    /* App background + global text (critical for Light Mode) */
    .stApp {{ background-color: {t['bg']} !important; color: {t['text']} !important; }}
    [data-testid="stMain"] {{ color: {t['text']} !important; }}
    [data-testid="stAppViewContainer"] {{ color: {t['text']} !important; }}
    .main .block-container {{ color: {t['text']} !important; }}
    [data-testid="stMarkdownContainer"],
    [data-testid="stMarkdownContainer"] p,
    [data-testid="stMarkdownContainer"] span {{ color: {t['text']} !important; }}
    /* Selectbox / widget labels */
    [data-testid="stSelectbox"] label,
    [data-testid="stTextInput"] label,
    [data-testid="stSelectbox"] div[data-baseweb="select"] span,
    [data-testid="stSelectbox"] div[data-baseweb="select"] div {{
        color: {t['text']} !important;
    }}
    /* Sidebar */
    section[data-testid="stSidebar"] {{
        background-color: {t['card_bg']} !important;
    }}
    section[data-testid="stSidebar"] * {{ color: {t['text']} !important; }}
    /* Report container */
    .rpt-container {{
        font-family: 'Segoe UI', Arial, sans-serif;
        font-size: {fs}px;
        color: {t['text']};
        max-width: 860px;
    }}
    .rpt-ticker {{
        color: {t['header']};
        font-size: {fs + 14}px;
        font-weight: 700;
        letter-spacing: 2px;
        margin-bottom: 2px;
    }}
    .rpt-company {{
        color: {t['subhead']};
        font-size: {fs + 2}px;
        font-weight: 600;
        margin-bottom: 4px;
    }}
    .rpt-desc {{
        color: {t['text']};
        font-size: {fs - 1}px;
        opacity: 0.75;
        margin-bottom: 10px;
        line-height: 1.5;
    }}
    .rpt-divider {{
        border: none;
        border-top: 2px solid {t['accent']};
        margin: 10px 0 18px 0;
    }}
    .rpt-thin-divider {{
        border: none;
        border-top: 1px solid {t['border']};
        margin: 4px 0 10px 0;
    }}
    .rpt-total-box {{
        margin: 0 0 22px 0;
        padding: 18px 24px;
        border: 2px solid {t['accent']};
        border-radius: 10px;
        background: {t['card_bg']};
        display: flex;
        align-items: center;
        gap: 16px;
    }}
    .rpt-total-lbl {{
        color: {t['text']};
        font-size: {fs + 1}px;
        font-weight: 700;
    }}
    .rpt-total-profit {{
        color: {t['profit']};
        font-size: {fs + 5}px;
        font-weight: 800;
    }}
    .rpt-total-loss {{
        color: {t['loss']};
        font-size: {fs + 5}px;
        font-weight: 800;
    }}
    .rpt-txn-hdr {{
        color: {t['subhead']};
        font-size: {fs + 1}px;
        font-weight: 700;
        margin: 18px 0 8px 0;
        padding-bottom: 4px;
        border-bottom: 1px solid {t['border']};
    }}
    .rpt-row {{
        display: flex;
        padding: 5px 0;
        border-bottom: 1px solid rgba(255,255,255,0.06);
        align-items: center;
    }}
    .rpt-lbl {{
        color: {t['text']};
        width: 220px;
        opacity: 0.8;
        font-size: {fs}px;
    }}
    .rpt-val {{
        color: {t['text']};
        font-weight: 600;
        font-size: {fs}px;
    }}
    .rpt-profit {{
        color: {t['profit']};
        font-weight: 700;
        font-size: {fs}px;
    }}
    .rpt-loss {{
        color: {t['loss']};
        font-weight: 700;
        font-size: {fs}px;
    }}
    /* Streamlit widget overrides */
    div[data-testid="stSelectbox"] label,
    div[data-testid="stTextInput"] label {{
        color: {t['subhead']} !important;
        font-weight: 600;
    }}
    div[data-testid="metric-container"] {{
        background: {t['card_bg']};
        border: 1px solid {t['border']};
        border-radius: 8px;
        padding: 10px 14px;
    }}
    div[data-testid="metric-container"] label {{
        color: {t['subhead']} !important;
    }}
    div[data-testid="metric-container"] div[data-testid="stMetricValue"] {{
        color: {t['header']} !important;
        font-weight: 700;
    }}
    </style>
    """, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# Report HTML builder
# ─────────────────────────────────────────────────────────────────────────────

def _build_report_html(
    portfolio: dict,
    ticker: str,
    live_price,
    yf_company: str,
    company_desc: str = "",
) -> str:
    data         = portfolio[ticker]
    transactions = data["transactions"]
    is_cad       = data["is_canadian"]

    # Strip Excel formula errors from company name
    excel_company = data["company_name"]
    if not excel_company or str(excel_company).strip().startswith("#"):
        excel_company = ""
    company = yf_company or excel_company or ticker

    def field_row(label: str, value: str, css_class: str = "rpt-val") -> str:
        return (
            f'<div class="rpt-row">'
            f'<span class="rpt-lbl">{label}</span>'
            f'<span class="{css_class}">{value}</span>'
            f'</div>'
        )

    # Pre-calculate total P&L
    total_profit = 0.0
    pnl_per_txn  = []
    for txn in transactions:
        try:
            cur  = live_price if live_price is not None else float(txn["current_price"])
            pur  = float(txn["purchase_price"])
            shs  = float(txn["shares"])
            pf   = (cur - pur) * shs
            pp   = ((cur - pur) / pur * 100) if pur != 0 else 0.0
            total_profit += pf
        except (TypeError, ValueError):
            pf, pp, cur = None, None, None
        pnl_per_txn.append((pf, pp, cur))

    html = ['<div class="rpt-container">']
    html.append(f'<div class="rpt-ticker">{ticker}</div>')
    if company:
        html.append(f'<div class="rpt-company">{company}</div>')
    # Company / ETF description (from stock_viewer_web pattern)
    if company_desc:
        html.append(f'<div class="rpt-desc">{company_desc}</div>')
    html.append('<hr class="rpt-divider">')

    # Total box at top
    total_class = "rpt-total-profit" if total_profit >= 0 else "rpt-total-loss"
    total_str   = fmt_currency(total_profit, is_cad, show_sign=True)
    html.append(
        f'<div class="rpt-total-box">'
        f'<span class="rpt-total-lbl">TOTAL PROFIT / LOSS:</span>'
        f'<span class="{total_class}">{total_str}</span>'
        f'</div>'
    )

    for idx, (txn, (p_float, p_pct, cur_used)) in enumerate(
        zip(transactions, pnl_per_txn), 1
    ):
        acct     = txn.get("account") or ""
        acct_str = (
            f" &nbsp;·&nbsp; <span style='opacity:0.7;font-weight:400'>{acct}</span>"
            if acct else ""
        )
        html.append(f'<div class="rpt-txn-hdr">Transaction #{idx}{acct_str}</div>')

        cur_display = cur_used if cur_used is not None else txn["current_price"]
        html.append(field_row("Current Price:",    fmt_currency(cur_display,           is_cad)))
        html.append(field_row("Purchase Price:",   fmt_currency(txn["purchase_price"], is_cad)))
        html.append(field_row("Amount of Shares:", fmt_shares(txn["shares"])))

        pnl_class = "rpt-profit" if (p_float or 0) >= 0 else "rpt-loss"
        pnl_str   = fmt_currency(p_float, is_cad, show_sign=True)
        pct_str   = fmt_pct(p_pct)
        html.append(field_row(
            "Profit / Loss:",
            f"{pnl_str} &nbsp;<span style='opacity:0.75'>({pct_str})</span>",
            pnl_class,
        ))

        html.append(field_row("Purchase Date:", fmt_date(txn["purchase_date"])))
        html.append(field_row("Selling Date:",  fmt_date(txn["selling_date"])))

    html.append('</div>')
    return "\n".join(html)


# ─────────────────────────────────────────────────────────────────────────────
# Streamlit App
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    st.set_page_config(
        page_title="Gaboch Portfolio",
        page_icon="📊",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    # ── Determine local vs cloud FIRST (needed for theme persistence) ────────
    excel_path   = st.session_state.get("excel_path", DEFAULT_EXCEL)
    local_exists = os.path.exists(excel_path)

    _init_theme()
    theme = _get_theme()
    _persist_theme_if_changed(theme, local_exists)   # auto-save on every change
    _inject_css(theme)
    load_error   = None
    portfolio: dict = {}

    if local_exists:
        try:
            portfolio = load_portfolio(excel_path)
        except Exception as exc:
            load_error = str(exc)
        if not load_error:
            _auto_push_if_changed(excel_path)   # silent sync to GitHub
    else:
        # Cloud deployment — try GitHub first, then fall back to session upload
        portfolio, gh_err = load_portfolio_github()
        if gh_err and gh_err != "not_found":
            load_error = gh_err
        if not portfolio and "excel_bytes" in st.session_state:
            try:
                portfolio = load_portfolio_bytes(st.session_state["excel_bytes"])
            except Exception as exc:
                load_error = str(exc)

    # ── Sidebar ───────────────────────────────────────────────────────────────
    with st.sidebar:
        st.markdown(
            f"<h2 style='color:{theme['header']};margin-bottom:0;padding-bottom:0'>"
            f"📊 Gaboch Portfolio</h2>",
            unsafe_allow_html=True,
        )
        st.markdown("---")

        if local_exists:
            # Laptop: show path input
            new_path = st.text_input("📁 Excel File Path:", value=excel_path)
            if new_path != excel_path:
                st.session_state.excel_path = new_path
                load_portfolio.clear()
                st.rerun()
        else:
            # Cloud: auto-loaded from GitHub — only show uploader if file not found
            _, gh_err2 = load_portfolio_github()
            if gh_err2 == "not_found":
                st.markdown(
                    f"<span style='color:{theme['subhead']};font-weight:600'>"
                    f"📁 Upload once — saves to GitHub automatically:</span>",
                    unsafe_allow_html=True,
                )
                uploaded = st.file_uploader(
                    "Gaboch_portfolio.xlsx", type=["xlsx", "xlsm"],
                    label_visibility="collapsed",
                )
                if uploaded is not None:
                    raw = uploaded.read()
                    st.session_state["excel_bytes"] = raw
                    load_portfolio_bytes.clear()
                    with st.spinner("Saving to GitHub for auto-load next time…"):
                        ok, push_err = push_excel_to_github(raw)
                    if ok:
                        load_portfolio_github.clear()
                        st.success("✓ Saved! Future visits load automatically.")
                    else:
                        st.error(push_err)
                    st.rerun()
            else:
                st.caption(f"✅ Auto-loaded from GitHub")

        if st.button("🔄 Refresh", use_container_width=True):
            load_portfolio.clear()
            load_portfolio_bytes.clear()
            load_portfolio_github.clear()
            fetch_yf_info.clear()
            st.session_state.pop("_excel_hash", None)   # force re-push on next load
            st.rerun()
        if local_exists:
            if st.session_state.get("_excel_hash"):
                st.caption("✅ Synced to GitHub")

        if load_error:
            st.error(f"⚠️ Load error:\n{load_error}")

        st.markdown("---")

        # Ticker selector
        tickers = sorted(portfolio.keys()) if portfolio else []
        if tickers:
            selected_ticker: str | None = st.selectbox(
                "📈 Select Ticker:", tickers,
                help="Only active (not sold) tickers are listed.",
            )
        else:
            selected_ticker = None
            if not load_error:
                st.info("No active tickers found.")

        st.markdown("---")

        # ── Customize Layout ──────────────────────────────────────────────────
        with st.expander("⚙️  Customize Layout", expanded=False):
            preset_name = st.selectbox(
                "Theme Preset:", list(THEME_PRESETS.keys()),
                label_visibility="visible",
            )
            col_a, col_b = st.columns(2)
            with col_a:
                if st.button("✔ Apply Preset", use_container_width=True, key="btn_preset"):
                    _apply_preset(THEME_PRESETS[preset_name])
                    st.rerun()
            with col_b:
                if st.button("↺ Defaults", use_container_width=True, key="btn_reset"):
                    _apply_preset(DEFAULTS)
                    st.rerun()

            st.markdown("**Colors** *(pick to apply instantly)*")

            # Color pickers — auto-apply on change (stock_viewer_web pattern)
            new_profit = st.color_picker("Profit",     theme["profit"],  key="cp_profit")
            if new_profit != theme["profit"]:
                st.session_state.theme["profit"] = new_profit
                st.rerun()

            new_loss = st.color_picker("Loss",         theme["loss"],    key="cp_loss")
            if new_loss != theme["loss"]:
                st.session_state.theme["loss"] = new_loss
                st.rerun()

            new_header = st.color_picker("Header",     theme["header"],  key="cp_header")
            if new_header != theme["header"]:
                st.session_state.theme["header"] = new_header
                st.rerun()

            new_sub = st.color_picker("Sub-header",   theme["subhead"], key="cp_sub")
            if new_sub != theme["subhead"]:
                st.session_state.theme["subhead"] = new_sub
                st.rerun()

            new_accent = st.color_picker("Accent",     theme["accent"],  key="cp_accent")
            if new_accent != theme["accent"]:
                st.session_state.theme["accent"] = new_accent
                st.rerun()

            new_text = st.color_picker("Text",         theme["text"],    key="cp_text")
            if new_text != theme["text"]:
                st.session_state.theme["text"] = new_text
                st.rerun()

            new_sz = st.slider(
                "Font Size (px)", min_value=12, max_value=22,
                value=theme["font_size"], step=1, key="sl_font",
            )
            if new_sz != theme["font_size"]:
                st.session_state.theme["font_size"] = new_sz
                st.rerun()

    # ── Main content ──────────────────────────────────────────────────────────
    if load_error:
        st.error(f"**Failed to load portfolio file:** `{excel_path}`")
        st.code(load_error)
        st.info("Update the Excel file path in the sidebar.")
        return

    if not selected_ticker or not portfolio:
        st.markdown(
            f"<div style='color:{theme['subhead']};font-size:18px;margin-top:40px;"
            f"text-align:center'>"
            f"📈 Select a ticker from the sidebar to view transactions."
            f"</div>",
            unsafe_allow_html=True,
        )
        return

    data = portfolio[selected_ticker]

    # ── Live Yahoo Finance data ───────────────────────────────────────────────
    yf_ticker = data.get("yf_ticker", "") or ""
    if not yf_ticker:
        base = selected_ticker.split()[0] if " " in selected_ticker else selected_ticker
        if data.get("is_canadian") and not base.upper().endswith(".TO"):
            yf_ticker = base + ".TO"
        else:
            yf_ticker = base

    yf_data      = fetch_yf_info(yf_ticker)
    yf_company   = yf_data.get("company_name", "")
    company_desc = yf_data.get("company_desc", "")
    live_price   = yf_data.get("current_price")
    is_cad       = data["is_canadian"] or yf_data.get("is_canadian", False)

    txns = data["transactions"]
    sym  = "CAD$" if is_cad else "USD$"

    # Company name with #VALUE! filter
    excel_company = data["company_name"]
    if not excel_company or str(excel_company).strip().startswith("#"):
        excel_company = ""
    company = yf_company or excel_company or selected_ticker

    # ── Summary metrics ───────────────────────────────────────────────────────
    total_pnl = 0.0
    for txn in txns:
        try:
            cur = live_price if live_price is not None else float(txn["current_price"])
            pur = float(txn["purchase_price"])
            shs = float(txn["shares"])
            total_pnl += (cur - pur) * shs
        except (TypeError, ValueError):
            pass

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Ticker",       selected_ticker)
    col2.metric("Company",      company[:28] + "…" if len(company) > 28 else company)
    col3.metric("Transactions", len(txns))
    col4.metric("Currency",     sym)

    st.markdown("<div style='margin-top:6px'></div>", unsafe_allow_html=True)

    # ── Transaction report ────────────────────────────────────────────────────
    html = _build_report_html(
        portfolio, selected_ticker, live_price, yf_company, company_desc
    )
    st.markdown(html, unsafe_allow_html=True)


if __name__ == "__main__":
    main()
